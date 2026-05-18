from flask import (
    send_file,
    Flask,
    Response
)

from openpyxl import Workbook

from services.logger import logger

import sqlite3


DB_PATH: str = "data/inventory.db"


def register_export_routes(app: Flask) -> None:
    """
    Register export-related application routes.
    """

    @app.route("/export")
    def export_excel() -> Response:
        """
        Export application data to an Excel file.
        """

        workbook = Workbook()

        with sqlite3.connect(DB_PATH) as conn:

            cursor = conn.cursor()

            # Employees sheet
            employees_sheet = workbook.active

            employees_sheet.title = "Employees"

            employees_sheet.append([
                "ID",
                "First Name",
                "Last Name",
                "Department"
            ])

            cursor.execute("""
                SELECT
                    id,
                    first_name,
                    last_name,
                    department
                FROM employees
            """)

            for row in cursor.fetchall():
                employees_sheet.append(row)

            # Assets sheet
            assets_sheet = workbook.create_sheet(
                title="Assets"
            )

            assets_sheet.append([
                "ID",
                "Category",
                "Brand",
                "Serial Number",
                "Status",
                "Purchase Date"
            ])

            cursor.execute("""
                SELECT
                    id,
                    category,
                    brand,
                    serial_number,
                    status,
                    purchase_date
                FROM assets
            """)

            for row in cursor.fetchall():
                assets_sheet.append(row)

            # Assignments sheet
            assignments_sheet = workbook.create_sheet(
                title="Assignments"
            )

            assignments_sheet.append([
                "ID",
                "Employee",
                "Asset",
                "Assigned Date",
                "Returned Date",
                "Status"
            ])

            cursor.execute("""
                SELECT
                    assignments.id,
                    employees.first_name || ' ' ||
                    employees.last_name,
                    assets.category || ' - ' ||
                    assets.brand,
                    assignments.assigned_date,
                    assignments.returned_date,
                    assignments.status
                FROM assignments
                JOIN employees
                    ON assignments.employee_id = employees.id
                JOIN assets
                    ON assignments.asset_id = assets.id
            """)

            for row in cursor.fetchall():
                assignments_sheet.append(row)

        file_name = "inventory_export.xlsx"

        workbook.save(file_name)

        logger.info(
            "Excel export generated successfully"
        )

        return send_file(
            file_name,
            as_attachment=True
        )